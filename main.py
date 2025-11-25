import pandas as pd
from docxtpl import DocxTemplate
import os
import openpyxl
from docx2pdf import convert
import check_data


def create_errorsFile(errors):
    data_errors = pd.DataFrame(list(errors.items()), columns=['Num', 'Error'])
    data_errors.to_excel('errors.xlsx', index=False)


def read_anketa(file) -> tuple:

    context = None
    flag = True
    etalon_true = ['Соответствует', 'Частично соответствует', 'Не соответствует']
    path = 'Raw_2025' + r'\\' + file                             # Raw
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet1 = wb['Форма']
    sheets = wb.sheetnames[1:]
    sheets_kt = wb.sheetnames[-16:-13]
    sheets_st = wb.sheetnames[-13:-10]
    context = {'num_project': sheet1['C2'].value,
               'name_project': sheet1['C3'].value,
               'code_project': sheet1['C4'].value,
               'org_project': sheet1['C5'].value,
               'start_project': sheet1['C6'].value,
               'end_project': sheet1['C7'].value}
    temp_counter, list_results = check_data.check_types(sheet1)
    list_true = check_data.check_true_result(sheet1)
    conc = check_data.check_conc(sheet1)
    
    if temp_counter != len(list_results):
        if sheet1['C2'].value in errors:
            errors[sheet1['C2'].value].append('Проблемы в Видах и/или Типах результатов')
        else:
            errors[sheet1['C2'].value] = ['Проблемы в Видах и/или Типах результатов']
    elif temp_counter != len(list_true):
        if sheet1['C2'].value in errors:
            errors[sheet1['C2'].value].append('Проблема в соответствии результатов')
        else:
            errors[sheet1['C2'].value] = ['Проблема в соответствии результатов']
    elif len(conc) == 0:
        if sheet1['C2'].value in errors:
            errors[sheet1['C2'].value].append('')
        else:
            errors[sheet1['C2'].value] = ['Нет выводов']
    else:

        context[f'conc_project'] = f'{conc}'.replace('..', '.')
        for idx_res in range(len(list_results)):
            context[f'name_result_{idx_res + 1}'] = f'{list_results[idx_res][0]}'
            context[f'type_result_{idx_res + 1}'] = f'{list_results[idx_res][1]}'

            if context[f'type_result_{idx_res + 1}'] == 'Лекарственный препарат':
                temp_result = 'ЛП'
            elif context[f'type_result_{idx_res + 1}'] == 'Медицинское изделие':
                temp_result = 'МИ'
            elif context[f'type_result_{idx_res + 1}'] == 'Клинические рекомендации':
                temp_result = 'КР'
            else:
                temp_result = 'Иное'
            sheet_ugt = wb[f'Р{idx_res + 1}_УГТ_{temp_result}']
            sheet_doc = wb[f'Р{idx_res + 1}_ДОК_{temp_result}']
            if sheet1['C7'].value == 2024:
                sheet_nioktr1 = wb[f'Р{idx_res + 1}_НИОКТР_1']
                if check_data.check_nioktr(sheet_nioktr1):
                    project_status = check_data.check_nioktr(sheet_nioktr1).split()
                    project_full_status = ' '.join(project_status[1:])
                    context[f'project_status_{idx_res + 1}'] = project_full_status[0].upper() + project_full_status[1:] + ':'
                    if project_status[0] == 'Нет,':
                        df_new_results = check_data.create_nioktr1_table(file, f'Р{idx_res + 1}_НИОКТР_1')
                        new_results = df_new_results.iloc[:, 0].to_list()
                        context[f'results_project_{idx_res + 1}'] = new_results
                    else:
                        df_new_results = check_data.create_nioktr2_table(file, f'Р{idx_res + 1}_НИОКТР_2')
                        new_results = df_new_results.to_dict(orient='records')
                        context[f'results2_project_{idx_res + 1}'] = new_results
                else:
                    errors[sheet1['C2'].value] = ['Не выбран ответ по результатам для проекта 2024 года']
                # if check_data.check_nioktr(sheet_nioktr1).startswith('Да'):
                #     context['positive_nioktr'] = check_data.check_nioktr(sheet_nioktr1)
                #     sheet_nioktr2 = wb[f'Р{idx_res + 1}_НИОКТР_2']
                # else:

            if check_data.check_ugt(sheet_ugt):
                df_ugt = check_data.create_ugt_table(file, f'Р{idx_res + 1}_УГТ_{temp_result}')
                ugt_data = df_ugt.to_dict(orient='records')
                context[f'tasks_project_{idx_res + 1}'] = ugt_data

            else:
                if sheet1['C2'].value in errors:
                    errors[sheet1['C2'].value].append(f'Ошибка в УГТ для результата {idx_res + 1}')
                else:
                    errors[sheet1['C2'].value] = [f'Ошибка в УГТ для результата {idx_res + 1}']

            if check_data.check_doc(sheet_doc):
                df_doc = check_data.create_doc_table(file, f'Р{idx_res + 1}_ДОК_{temp_result}')
                docs_data = df_doc.to_dict(orient='records')
                context[f'docs_project_{idx_res + 1}'] = docs_data
            else:
                if sheet1['C2'].value in errors:
                    errors[sheet1['C2'].value].append(f'Ошибка в ДОК для результата {idx_res + 1}')
                else:
                    errors[sheet1['C2'].value] = [f'Ошибка в ДОК для результата {idx_res + 1}']

            if list_true[idx_res] not in etalon_true:
                if sheet1['C2'].value in errors:
                    errors[sheet1['C2'].value].append(f'Некорректный термин соответствия для результата {idx_res + 1}')
                else:
                    errors[sheet1['C2'].value] = [f'Некорректный термин соответствия для результата {idx_res + 1}']

            context[f'true_result_{idx_res + 1}'] = f'{list_true[idx_res]}'


            list_all_problems = check_data.check_problems(wb[sheets[idx_res]])
            list_all_kts = check_data.check_kt(wb[sheets_kt[idx_res]])
            list_all_sts = check_data.check_st(wb[sheets_st[idx_res]])

            for count_problems, problems in enumerate(list_all_problems):
                context[f'problem_{idx_res + 1}_{count_problems + 1}'] = problems[0]
                context[f'effect_{idx_res + 1}_{count_problems + 1}'] = problems[1]
                context[f'import_{idx_res + 1}_{count_problems + 1}'] = problems[2]

            for count_kt, kt in enumerate(list_all_kts):
                context[f'kt_{idx_res + 1}_{count_kt + 1}'] = kt

            for count_st, st in enumerate(list_all_sts):
                context[f'st_{idx_res + 1}_{count_st + 1}'] = st

            if len(list_all_problems) == 0:
                if sheet1['C2'].value in errors:
                    errors[sheet1['C2'].value].append('Не выбрано ни одной проблемы')
                else:
                    errors[sheet1['C2'].value] = ['Не выбрано ни одной проблемы']

    for key, value in errors.items():
        print(key, value, sep='\n')

    return len(list_results), context


def find_anketa(num_project):
    for address, dirs, files in os.walk('Raw_2025'):             # Raw
        for file in files:
            if 'KPM' in file:
                new_file = file.replace('KPM', 'КПМ')
            elif 'Lab' in file:
                new_file = file.replace('Lab', 'Лаб')
            else:
                break
            file_name = new_file[7:-5]
            if file_name == num_project:
                return read_anketa(file)


def create_pdf(file):
    pdf_file = f'{file[:-4]}pdf'
    convert('Doc_2025\\' + file, 'PDF_2025\\' + pdf_file)


def create_context(data):
    global errors
    errors = {}
    for key, values in data.items():
        try:
            count_results, context = find_anketa(key)
            context['zadanie'] = values[0]
            context['expert_project'] = values[1]
            if count_results > 0:
                doc = DocxTemplate(f'Template_result_{count_results}.docx')
                doc.render(context)
                doc_file = f'Zakluychenie {key}.docx'
                doc.save('Doc_2025\\' + doc_file)
                create_pdf(doc_file)
        except Exception:
            print(f'{key} File not created')
            # break
        finally:
            if len(errors) > 0:
                create_errorsFile(errors)


def create_dict_zadaniya() -> dict:
    dict_zadaniya = {}
    df_zadaniya = pd.read_excel('zadaniya_2025.xlsx', sheet_name='ФАРМА')
    for i in range(df_zadaniya['Number'].shape[0]):
        dict_zadaniya[df_zadaniya['Number'][i]] = [df_zadaniya['Zadanie'][i], df_zadaniya['Expert short'][i]]

    return dict_zadaniya


def main():
    create_context(create_dict_zadaniya())


if __name__ == '__main__':
    main()
